"""
Upload toan bo file xlsx trong folder vao hanja_tree_nodes
Bo qua tu trung lap (ON CONFLICT korean DO NOTHING)
"""
import os
import re
import requests
import openpyxl

FOLDER = r"C:\Users\hi\Desktop\Ebook\excel_output\1"
SUPABASE_URL = "https://dcjofhkdrgbrowabudyt.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRjam9maGtkcmdicm93YWJ1ZHl0Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjE1OTczNiwiZXhwIjoyMDkxNzM1NzM2fQ.T1_WxXzgB0LhFxcOvlqLyt_83rMOgmaQIuUO_4stPOE"
BATCH = 50

def get_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }

def get_root_char(hanja: str) -> str:
    for ch in hanja:
        if '\u4e00' <= ch <= '\u9fff':
            return ch
    return hanja[0] if hanja else ""

def clean_korean(text: str) -> str:
    # Xoa so o cuoi vi du: 가격03 -> 가격
    return re.sub(r'\d+$', '', str(text).strip())

def read_xlsx(filepath):
    rows = []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            obj = {headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row)}
            korean = clean_korean(obj.get("Tiếng Hàn", "") or obj.get("tiếng hàn", ""))
            hanja  = obj.get("Hán tự", "") or obj.get("hán tự", "")
            viet   = obj.get("Nghĩa tiếng Việt", "") or obj.get("nghĩa tiếng việt", "")
            if not korean or not viet:
                continue
            rows.append({
                "korean": korean,
                "hanja": hanja,
                "vietnamese": viet,
                "root_char": get_root_char(hanja) if hanja else korean[0],
                "difficulty": 2,
                "category": "Khác",
            })
    except Exception as e:
        print(f"  Loi doc file: {e}", flush=True)
    return rows

def insert_batch(rows):
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/hanja_tree_nodes",
        headers=get_headers(),
        json=rows,
        timeout=30
    )
    return r.status_code in (200, 201)

def main():
    files = sorted([f for f in os.listdir(FOLDER) if f.endswith(".xlsx")])
    print(f"Tim thay {len(files)} file xlsx", flush=True)

    total_rows = 0
    total_inserted = 0

    for fname in files:
        fpath = os.path.join(FOLDER, fname)
        rows = read_xlsx(fpath)
        total_rows += len(rows)

        inserted = 0
        for i in range(0, len(rows), BATCH):
            batch = rows[i:i+BATCH]
            if insert_batch(batch):
                inserted += len(batch)
        total_inserted += inserted
        print(f"  {fname}: {len(rows)} dong -> insert {inserted}", flush=True)

    print(f"\nHOAN THANH! Tong {total_rows} dong, insert {total_inserted}", flush=True)
    print("(Cac tu da co trong DB se bi bo qua tu dong)", flush=True)

if __name__ == "__main__":
    main()
